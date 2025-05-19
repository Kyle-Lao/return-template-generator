from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, date
from typing import Dict, List

def generate_return_template(
    insured_name: str,
    dob: str,
    carrier: str,
    le_months: int,
    le_report_date: str,
    death_benefit: float,
    investment: float,
    monthly_premiums: Dict[int, List[float]],
    output_filename: str
) -> str:
    dob = datetime.strptime(dob, "%Y-%m-%d")
    le_report_date = datetime.strptime(le_report_date, "%Y-%m-%d")
    today = date.today()

    elapsed_months = (today.year - le_report_date.year) * 12 + today.month - le_report_date.month
    remaining_le_months = max(le_months - elapsed_months, 0)
    remaining_le_years = (remaining_le_months + 11) // 12
    total_years = remaining_le_years + 3
    start_year = le_report_date.year
    age = int((le_report_date - dob).days / 365.25 + elapsed_months / 12)

    annual_premiums = {year: sum(months) for year, months in monthly_premiums.items()}

    wb = load_workbook("return_template_output.xlsx")
    ws = wb.active

    for _ in range(7, ws.max_row + 1):
        ws.delete_rows(7)

    ws["B1"] = insured_name
    ws["B2"] = f"AGE: {age}"
    ws["B3"] = f"CARRIER: {carrier}"
    ws["E2"] = f"{remaining_le_months} MONTHS"
    ws["E3"] = death_benefit
    ws["E4"] = investment

    cumulative = 0
    for i in range(total_years):
        year = start_year + i
        premium = annual_premiums.get(year, 0)
        cumulative += premium
        total_cost = investment + cumulative
        profit = death_benefit - total_cost
        simple_return = profit / total_cost if total_cost else 0
        acc_return = simple_return / (i + 1) if (i + 1) else 0
        marker = ""
        if i == remaining_le_years - 1:
            marker = "LE"
        elif i == remaining_le_years:
            marker = "LE+1"
        elif i == remaining_le_years + 1:
            marker = "LE+2"
        elif i == remaining_le_years + 2:
            marker = "LE+3"

        row = [
            year,
            premium,
            cumulative,
            total_cost,
            profit,
            simple_return,
            acc_return,
            marker
        ]
        ws.append(row)

        if i == remaining_le_years - 1:
            for col in range(2, 9):
                cell = ws.cell(row=6 + i + 1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for row in range(7, 7 + total_years):
        for col in range(2, 6):
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
        for col in range(6, 8):
            ws.cell(row=row, column=col).number_format = '0.00%'

    ref_style = ws["B6"]._style
    for row in range(7, 7 + total_years):
        ws.cell(row=row, column=1)._style = ref_style

    if ws["H6"].value == "LE Marker":
        ws["H6"].value = ""

    wb.save(output_filename)
    return output_filename
